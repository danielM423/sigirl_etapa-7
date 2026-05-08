
import os
import sys
import django

# Agrega la raíz del proyecto al sys.path para que Python encuentre 'sigirl'
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

# Agrega la raíz del proyecto al sys.path
sys.path.append(os.path.dirname(__file__))

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'sigirl.settings')
django.setup()

from django.contrib.auth.models import User
import openpyxl
from datetime import date
from inventario.models import Practica
from django.contrib.auth.models import User

# Obtén un usuario existente (ajusta el filtro si es necesario)
instructor = User.objects.first()  # O User.objects.get(username='tu_usuario')

# Crear la práctica con los campos correctos
p = Practica.objects.create(
    grupo='12/GID',
    especialidad='Química',
    fecha='2026-05-07',  # O date(2026, 5, 7)
    hora='08:00',
    instructor=instructor,  # Puede ser None si permites nulo
    nombre_practica='Práctica de Laboratorio'
)

print(f"✅ Práctica creada: {p.nombre_practica} - {p.fecha}")

# Ruta a la plantilla original
PLANTILLA = 'plantilla_practica.xlsx'
# Ruta de salida
SALIDA = 'practica_generada.xlsx'

# Datos de ejemplo (puedes reemplazar por datos de tu base de datos)
encabezado = {
    'grupo': '12/GID',
    'especialidad': 'QUÍMICA APLICADA A LA INDUSTRIA',
    'fecha': '2026-03-24',
    'hora': '6:00',
    'instructor': 'Giovanny Fonse',
    'nombre_practica': 'DETERMINACIÓN DE K, D, Mg. DETERMINACIÓN DE LA DOSIS ÓPTIMA DE COAGULANTE',
}

reactivos = [
    {'nombre': 'Ácido clorhídrico', 'cantidad': 1, 'unidad': 'L'},
    {'nombre': 'Sulfato de aluminio', 'cantidad': 2, 'unidad': 'g'},
    # ...más reactivos...
]

materiales = [
    {'nombre': 'Vaso de precipitados 100 mL', 'cantidad_por_grupo': 2, 'cantidad_total': 6},
    # ...más materiales...
]

equipos = [
    {'nombre': 'Balanza analítica', 'cantidad': 1},
    # ...más equipos...
]

def llenar_plantilla():
    wb = openpyxl.load_workbook(PLANTILLA)
    ws = wb.active

    # === Encabezado ===
    ws['D2'] = encabezado['grupo']
    ws['G2'] = encabezado['especialidad']
    ws['D3'] = encabezado['fecha']
    ws['G3'] = encabezado['hora']
    ws['G4'] = encabezado['instructor']
    ws['B5'] = encabezado['nombre_practica']

    # === Reactivos (ajusta las filas/columnas según tu plantilla) ===
    fila_reactivos = 10  # Fila donde empiezan los reactivos
    for i, r in enumerate(reactivos):
        ws[f'B{fila_reactivos + i}'] = r['nombre']
        ws[f'C{fila_reactivos + i}'] = r['cantidad']
        ws[f'D{fila_reactivos + i}'] = r['unidad']

    # === Materiales ===
    fila_materiales = 28  # Fila donde empiezan los materiales
    for i, m in enumerate(materiales):
        ws[f'B{fila_materiales + i}'] = m['nombre']
        ws[f'C{fila_materiales + i}'] = m['cantidad_por_grupo']
        ws[f'D{fila_materiales + i}'] = m['cantidad_total']

    # === Equipos ===
    fila_equipos = 41  # Fila donde empiezan los equipos
    for i, eq in enumerate(equipos):
        ws[f'A{fila_equipos + i}'] = eq['nombre']
        ws[f'B{fila_equipos + i}'] = eq['cantidad']

    wb.save(SALIDA)
    print(f'Archivo generado: {SALIDA}')

if __name__ == '__main__':
    llenar_plantilla()