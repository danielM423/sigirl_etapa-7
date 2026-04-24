import csv
import re
import unicodedata
from datetime import datetime
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError

from inventario.models import Categoria, Producto


SKIP_NAME_PATTERNS = ("documento", "inventario")
INVALID_NAME_MARKERS = {
    "cantidad inicial",
    "cantidad inicial g",
    "existencia",
    "existencia g",
    "existencia g ml",
    "densidad g ml",
    "categoria",
    "nombre",
    "total",
    "fecha",
    "reviso",
    "reactivo",
    "enero",
    "febrero",
    "marzo",
    "abril",
    "mayo",
    "junio",
    "julio",
    "agosto",
    "septiembre",
    "octubre",
    "noviembre",
    "diciembre",
}


def _normalize_text(value):
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return " ".join(text.split())


def _to_int(value, default_value):
    if value in (None, ""):
        return default_value
    try:
        if isinstance(value, str):
            value = value.replace(",", ".")
        return int(float(value))
    except (TypeError, ValueError):
        return default_value


def _to_date(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()

    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _guess_column(key_map, candidates):
    for candidate in candidates:
        if candidate in key_map:
            return key_map[candidate]
    return None


def _should_skip_name(value):
    if value in (None, ""):
        return True
    if isinstance(value, datetime):
        return True

    text = str(value).strip()
    normalized = _normalize_text(text)
    if not normalized:
        return True
    if normalized in INVALID_NAME_MARKERS:
        return True
    if re.fullmatch(r"\d{4}", normalized):
        return True
    if _to_date(text) is not None:
        return True
    return False


def _read_csv(file_path):
    with file_path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        rows = list(reader)
        headers = reader.fieldnames or []
    return headers, rows


def _read_xlsx(file_path):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise CommandError(
            "Falta openpyxl. Instala dependencias con: pip install -r sigirl/sigirl/requirements.txt"
        ) from exc

    workbook = load_workbook(filename=file_path, data_only=True)
    sheet = workbook.active

    raw_rows = []
    for row in sheet.iter_rows(values_only=True):
        raw_rows.append([cell if cell is not None else "" for cell in row])

    reactivo_name = None
    cantidad_existencia = None
    for row in raw_rows[:10]:
        normalized_row = [_normalize_text(cell) for cell in row]
        if "reactivo" in normalized_row:
            reactivo_index = normalized_row.index("reactivo")
            for cell in row[reactivo_index + 1 :]:
                if not _should_skip_name(cell):
                    reactivo_name = str(cell).strip()
                    break

    for row in raw_rows[:10]:
        normalized_row = [_normalize_text(cell) for cell in row]
        if any(cell.startswith("existencia") for cell in normalized_row):
            numeric_values = []
            for cell in row:
                if isinstance(cell, (int, float)) and cell > 0:
                    numeric_values.append(cell)
            if numeric_values:
                cantidad_existencia = int(max(numeric_values))
                break

    if reactivo_name:
        return ["nombre", "cantidad"], [
            {
                "nombre": reactivo_name,
                "cantidad": cantidad_existencia or 0,
            }
        ]

    header_row_index = None
    for idx, row in enumerate(raw_rows[:12]):
        non_empty = [cell for cell in row if str(cell).strip()]
        if len(non_empty) >= 2:
            header_row_index = idx
            break

    if header_row_index is None:
        return [], []

    headers = [str(cell).strip() for cell in raw_rows[header_row_index]]
    rows = []
    for row in raw_rows[header_row_index + 1 :]:
        if not any(str(cell).strip() for cell in row):
            continue
        row_dict = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            row_dict[header] = row[idx] if idx < len(row) else ""
        rows.append(row_dict)

    return headers, rows


class Command(BaseCommand):
    help = "Importa reactivos/productos por lotes desde una carpeta con multiples Excel o CSV"

    def add_arguments(self, parser):
        parser.add_argument("--dir", required=True, help="Carpeta con archivos .xlsx, .xls o .csv")
        parser.add_argument(
            "--tipo-default",
            default="reactivo",
            choices=["reactivo", "insumo", "equipo"],
            help="Tipo por defecto cuando no exista columna de tipo",
        )
        parser.add_argument(
            "--minimo-default",
            type=int,
            default=1,
            help="Minimo por defecto cuando no exista columna de minimo",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Analiza y muestra resultados sin guardar en base de datos",
        )

    def handle(self, *args, **options):
        source_dir = Path(options["dir"]).expanduser().resolve()
        if not source_dir.exists() or not source_dir.is_dir():
            raise CommandError(f"Carpeta invalida: {source_dir}")

        supported_files = sorted(
            [
                file_path
                for file_path in source_dir.iterdir()
                if file_path.suffix.lower() in {".xlsx", ".xls", ".csv"}
                and not any(pattern in file_path.name.lower() for pattern in SKIP_NAME_PATTERNS)
            ]
        )

        if not supported_files:
            raise CommandError("No se encontraron archivos .xlsx, .xls o .csv validos en la carpeta")

        created_count = 0
        updated_count = 0
        skipped_count = 0
        errors = []

        self.stdout.write(f"Procesando {len(supported_files)} archivos en: {source_dir}")

        for file_path in supported_files:
            self.stdout.write(f"\nArchivo: {file_path.name}")

            try:
                if file_path.suffix.lower() == ".csv":
                    headers, rows = _read_csv(file_path)
                else:
                    headers, rows = _read_xlsx(file_path)
            except Exception as exc:
                errors.append(f"{file_path.name}: {exc}")
                self.stdout.write(self.style.ERROR(f"  Error leyendo archivo: {exc}"))
                continue

            if not rows:
                self.stdout.write(self.style.WARNING("  Sin filas utiles, se omite"))
                skipped_count += 1
                continue

            key_map = {_normalize_text(h): h for h in headers if h}
            nombre_col = _guess_column(key_map, ["nombre", "producto", "reactivo", "descripcion", "material", "item"])
            cantidad_col = _guess_column(key_map, ["cantidad", "stock", "existencia", "existencias"])
            minimo_col = _guess_column(key_map, ["minimo", "stock minimo", "min", "min stock"])
            ubicacion_col = _guess_column(key_map, ["ubicacion", "locacion", "almacen", "estante"])
            venc_col = _guess_column(key_map, ["fecha vencimiento", "vencimiento", "fecha de vencimiento", "vence"])
            tipo_col = _guess_column(key_map, ["tipo", "clase"])
            categoria_col = _guess_column(key_map, ["categoria", "familia", "grupo"])

            if not nombre_col:
                nombre_col = next(iter(key_map.values()), None)

            if not nombre_col:
                self.stdout.write(self.style.WARNING("  No se detecto columna de nombre, se omite"))
                skipped_count += 1
                continue

            category_from_file = file_path.stem.strip()
            local_created = 0
            local_updated = 0
            local_skipped = 0

            for row in rows:
                raw_name = row.get(nombre_col, "")
                if _should_skip_name(raw_name):
                    local_skipped += 1
                    continue
                nombre = str(raw_name).strip()

                categoria_nombre = str(row.get(categoria_col, "")).strip() if categoria_col else ""
                if not categoria_nombre:
                    categoria_nombre = category_from_file

                categoria_obj, _ = Categoria.objects.get_or_create(nombre=categoria_nombre)
                raw_tipo = str(row.get(tipo_col, "")).strip().lower() if tipo_col else ""
                tipo = raw_tipo if raw_tipo in {"reactivo", "insumo", "equipo"} else options["tipo_default"]
                cantidad = _to_int(row.get(cantidad_col, "") if cantidad_col else "", 0)
                minimo = _to_int(row.get(minimo_col, "") if minimo_col else "", options["minimo_default"])
                ubicacion = str(row.get(ubicacion_col, "")).strip() if ubicacion_col else ""
                fecha_venc = _to_date(row.get(venc_col, "") if venc_col else "")

                existing = Producto.objects.filter(nombre=nombre, categoria=categoria_obj).first()
                if existing:
                    existing.tipo = tipo
                    existing.cantidad = cantidad
                    existing.minimo = minimo
                    existing.ubicacion = ubicacion or None
                    existing.fecha_vencimiento = fecha_venc
                    if not options["dry_run"]:
                        existing.save()
                    local_updated += 1
                else:
                    if not options["dry_run"]:
                        Producto.objects.create(
                            nombre=nombre,
                            tipo=tipo,
                            categoria=categoria_obj,
                            cantidad=cantidad,
                            minimo=minimo,
                            ubicacion=ubicacion or None,
                            fecha_vencimiento=fecha_venc,
                        )
                    local_created += 1

            created_count += local_created
            updated_count += local_updated
            skipped_count += local_skipped
            self.stdout.write(self.style.SUCCESS(f"  OK -> creados: {local_created}, actualizados: {local_updated}, omitidos: {local_skipped}"))

        self.stdout.write("\nResumen final")
        self.stdout.write(f"- Creados: {created_count}")
        self.stdout.write(f"- Actualizados: {updated_count}")
        self.stdout.write(f"- Omitidos: {skipped_count}")
        self.stdout.write(f"- Archivos con error: {len(errors)}")

        if errors:
            self.stdout.write("\nDetalle de errores:")
            for err in errors:
                self.stdout.write(f"  - {err}")

        if options["dry_run"]:
            self.stdout.write(self.style.WARNING("\nDry-run activo: no se guardaron cambios."))