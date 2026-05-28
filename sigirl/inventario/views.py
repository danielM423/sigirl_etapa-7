# --- IMPORTS PRINCIPALES ---
from rest_framework import viewsets, permissions
# === RF-034: Historial de pedidos ===
from .models import Competencia, PedidoHistorial, PDFDocumento, Asistencia, ListadoDiario, Programa
from .serializers import CompetenciaSerializer, PedidoHistorialSerializer, PDFDocumentoSerializer, AsistenciaSerializer, ListadoDiarioSerializer, ProgramaSerializer
from rest_framework.parsers import MultiPartParser, FormParser

class PedidoHistorialViewSet(viewsets.ModelViewSet):
    queryset = PedidoHistorial.objects.none()
    serializer_class = PedidoHistorialSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        queryset = PedidoHistorial.objects.filter(usuario_modificador=user)
        fecha = self.request.query_params.get('fecha')
        if fecha:
            queryset = queryset.filter(fecha__date=fecha)
        return queryset

class PDFDocumentoViewSet(viewsets.ModelViewSet):
    queryset = PDFDocumento.objects.all()
    serializer_class = PDFDocumentoSerializer
    parser_classes = (MultiPartParser, FormParser)
    permission_classes = [permissions.IsAuthenticated]

    def perform_create(self, serializer):
        # Al guardar un PDF, registrar el usuario autenticado
        serializer.save(usuario=self.request.user)

class AsistenciaViewSet(viewsets.ModelViewSet):
    queryset = Asistencia.objects.all()
    serializer_class = AsistenciaSerializer
    permission_classes = [permissions.IsAuthenticated]

    def perform_create(self, serializer):
        # Registrar automáticamente el usuario autenticado
        serializer.save(usuario=self.request.user)

class ListadoDiarioViewSet(viewsets.ModelViewSet):
    queryset = ListadoDiario.objects.all()
    serializer_class = ListadoDiarioSerializer
    permission_classes = [permissions.IsAuthenticated]

    def perform_create(self, serializer):
        # Registrar automáticamente el usuario autenticado como creador
        serializer.save(creado_por=self.request.user)
from django.contrib.auth import get_user_model
User = get_user_model()
# Endpoint para obtener instructores (todos los usuarios)
from rest_framework.decorators import api_view
@api_view(['GET'])
def instructores_list(request):
    users = User.objects.all().values('id', 'username', 'first_name', 'last_name')
    return Response(list(users))
from django.db.models import Sum
from rest_framework.response import Response
# Endpoint para top de reactivos más usados
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny
@api_view(['GET'])
@permission_classes([AllowAny])  # Cambia a IsAuthenticated si lo necesitas
def top_reactivos_usados(request):
    top = (
        PracticaReactivo.objects
        .values('reactivo__nombre')
        .annotate(total=Sum('cantidad'))
        .order_by('-total')[:10]
    )
    return Response(list(top))
from django.db.models import Prefetch
# =====================
# PRÁCTICAS - Excel listado completo
# =====================
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework import status
from .models import Practica, PracticaReactivo, Producto
from .serializers import ProductoSerializer
from rest_framework.decorators import api_view, permission_classes
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def download_practicas_excel(request):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Prácticas'

    headers = [
        'ID', 'Ficha', 'Nombre', 'Fecha', 'Grupos', 'Instructor', 'Estado',
        'Reactivos', 'Equipos'
    ]
    sheet.append(headers)

    practicas = Practica.objects.select_related('instructor').prefetch_related(
        Prefetch('reactivos', to_attr='reactivos_list'),
        Prefetch('equipos', to_attr='equipos_list')
    ).order_by('-fecha')

    for p in practicas:
        reactivos_str = '\n'.join([
            f"{r.reactivo.nombre} ({r.cantidad} {r.unidad.simbolo}){' [Sensible]' if r.es_sensible else ''}"
            for r in getattr(p, 'reactivos_list', [])
        ])
        equipos_str = '\n'.join([
            f"{e.equipo.nombre} ({e.tiempo_uso_min} min, desgaste: {e.desgaste_estimado}, mantenimiento: {'Sí' if e.mantenimiento_requerido else 'No'})"
            for e in getattr(p, 'equipos_list', [])
        ])
        sheet.append([
            p.id,
            p.ficha,
            p.nombre,
            p.fecha.strftime('%Y-%m-%d'),
            p.grupos_trabajo,
            p.instructor.get_full_name() or p.instructor.username,
            p.estado,
            reactivos_str,
            equipos_str,
        ])

    # Estilos de encabezado
    header_fill = PatternFill(start_color='1FA971', end_color='1FA971', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    for column_index in range(1, len(headers) + 1):
        cell = sheet.cell(row=1, column=column_index)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        sheet.column_dimensions[chr(64 + column_index)].width = 22

    # Ajuste de alineación para celdas de datos
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="practicas_sigirl.xlsx"'
    return response

# === NUEVO ENDPOINT: Inventario por práctica abierta solo para instructores ===
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def inventario_practicas_abiertas_instructor(request):
    user = request.user
    # Ahora admin y jefe pueden ver el inventario de todas las prácticas abiertas
    if user.is_superuser or user.is_staff:
        practicas = Practica.objects.filter(estado__in=['aprobada', 'pendiente'])
    else:
        practicas = Practica.objects.filter(instructor=user, estado__in=['aprobada', 'pendiente'])
        if not practicas.exists():
            return Response({'detail': 'No tienes prácticas abiertas asignadas.'}, status=404)
    reactivos = PracticaReactivo.objects.filter(practica__in=practicas).select_related('reactivo')
    productos = [pr.reactivo for pr in reactivos]
    productos_unicos = {p.id: p for p in productos}.values()
    data = ProductoSerializer(productos_unicos, many=True).data
    return Response(data)
from rest_framework import viewsets, permissions
from .models import UnidadMedida, Practica, PracticaReactivo, PracticaEquipo, Producto
from .serializers import UnidadMedidaSerializer, PracticaSerializer, PracticaReactivoSerializer, PracticaEquipoSerializer, ProductoSerializer

# API REST para Unidades de Medida
class UnidadMedidaViewSet(viewsets.ModelViewSet):
    queryset = UnidadMedida.objects.all()
    serializer_class = UnidadMedidaSerializer

# API REST para Prácticas
class PracticaViewSet(viewsets.ModelViewSet):
    queryset = Practica.objects.all()
    serializer_class = PracticaSerializer
    permission_classes = [permissions.AllowAny]  # Solo para pruebas, luego volver a IsAuthenticated

    from rest_framework.decorators import action
    from rest_framework.response import Response
    from rest_framework import status

    @action(detail=True, methods=['post'], url_path='aprobar')
    def aprobar(self, request, pk=None):
        from django.db import transaction
        practica = self.get_object()
        user = request.user
        role = 'jefe' if user.is_superuser else 'admin' if user.is_staff else 'usuario'
        if role not in ['jefe', 'admin']:
            return self._forbidden_response()
        if practica.estado not in ['pendiente', 'en aprobación']:
            return Response({'error': 'La práctica no puede ser aprobada en su estado actual.'}, status=status.HTTP_400_BAD_REQUEST)
        if role == 'jefe':
            practica.jefe_aprobado = True
        if role == 'admin':
            practica.admin_aprobado = True

        # Si requiere doble aprobación, ambas deben estar en True para aprobar
        aprobar_definitivo = False
        if practica.requiere_doble_aprobacion:
            if practica.jefe_aprobado and practica.admin_aprobado:
                aprobar_definitivo = True
                practica.estado = 'aprobada'
            else:
                practica.estado = 'en aprobación'
        else:
            aprobar_definitivo = True
            practica.estado = 'aprobada'


        # Validación y descuento de stock solo si se aprueba definitivamente
        if aprobar_definitivo:
            reactivos = practica.reactivos.all()
            errores = []
            errores_sensibles = []
            # Validar stock suficiente y reactivos sensibles
            for pr in reactivos:
                producto = pr.reactivo
                cantidad_necesaria = pr.cantidad
                if producto.cantidad < cantidad_necesaria:
                    errores.append(f"Stock insuficiente para {producto.nombre} (disponible: {producto.cantidad}, requerido: {cantidad_necesaria})")
                if pr.es_sensible and producto.cantidad < cantidad_necesaria:
                    errores_sensibles.append(f"Reactivo sensible sin stock suficiente: {producto.nombre}")
            # Si hay error de reactivo sensible, bloquear y mostrar mensaje especial
            if errores_sensibles:
                return Response({'error': 'No se puede aprobar la práctica por reactivos sensibles sin stock suficiente', 'detalles': errores_sensibles}, status=status.HTTP_400_BAD_REQUEST)
            # Si hay error de stock general, bloquear
            if errores:
                return Response({'error': 'No se puede aprobar la práctica', 'detalles': errores}, status=status.HTTP_400_BAD_REQUEST)
            # Descontar stock
            with transaction.atomic():
                for pr in reactivos:
                    producto = pr.reactivo
                    producto.cantidad -= pr.cantidad
                    producto.save()
                practica.save()
        else:
            practica.save()

        return Response({'status': practica.estado, 'practica': PracticaSerializer(practica).data})

    @action(detail=True, methods=['post'], url_path='rechazar')
    def rechazar(self, request, pk=None):
        practica = self.get_object()
        user = request.user
        role = 'jefe' if user.is_superuser else 'admin' if user.is_staff else 'usuario'
        if role not in ['jefe', 'admin']:
            return self._forbidden_response()
        if practica.estado not in ['pendiente', 'en aprobación']:
            return Response({'error': 'La práctica no puede ser rechazada en su estado actual.'}, status=status.HTTP_400_BAD_REQUEST)
        practica.estado = 'rechazada'
        practica.save()
        return Response({'status': 'rechazada', 'practica': PracticaSerializer(practica).data})

    def _forbidden_response(self):
        from rest_framework.response import Response
        from rest_framework import status
        return Response({'error': 'No tienes permiso para esta acción.'}, status=status.HTTP_403_FORBIDDEN)

# API REST para Reactivos (productos tipo reactivo)
class ReactivoViewSet(viewsets.ModelViewSet):
    queryset = Producto.objects.filter(tipo='reactivo')
    serializer_class = ProductoSerializer

# API REST para Equipos (productos tipo equipo)
class EquipoViewSet(viewsets.ModelViewSet):
    queryset = Producto.objects.filter(tipo='equipo')
    serializer_class = ProductoSerializer

# API REST para PracticaReactivo y PracticaEquipo (opcional, si quieres exponerlos)
class PracticaReactivoViewSet(viewsets.ModelViewSet):
    queryset = PracticaReactivo.objects.all()
    serializer_class = PracticaReactivoSerializer

class PracticaEquipoViewSet(viewsets.ModelViewSet):
    queryset = PracticaEquipo.objects.all()
    serializer_class = PracticaEquipoSerializer
from datetime import date, timedelta
from io import BytesIO
import secrets

from django.conf import settings
from django.contrib.auth.models import User
from django.contrib.auth.tokens import default_token_generator
from django.contrib.auth.hashers import make_password, check_password
from django.core.exceptions import ValidationError
from django.core.mail import send_mail
from django.core.validators import validate_email
from django.http import HttpResponse
from django.utils import timezone
from django.utils.encoding import force_bytes, force_str
from django.utils.http import urlsafe_base64_decode, urlsafe_base64_encode

from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework_simplejwt.views import TokenObtainPairView, TokenRefreshView
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .models import Alerta, Categoria, Producto, Movimiento, Pedido, UserProfile
from .serializers import (
    AlertaSerializer,
    CategoriaSerializer,
    ProductoSerializer,
    MovimientoSerializer,
    PedidoSerializer,
    UserManagementSerializer,
)


class IsStaffOrSuperuser(permissions.BasePermission):
    def has_permission(self, request, view):
        return bool(request.user and request.user.is_authenticated and (request.user.is_staff or request.user.is_superuser))


class IsInventoryManagerOrReadOnly(permissions.BasePermission):
    def has_permission(self, request, view):
        if not request.user or not request.user.is_authenticated:
            return False
        if request.method in permissions.SAFE_METHODS:
            return True
        return request.user.is_staff or request.user.is_superuser


def _is_email_domain_reachable(domain: str) -> bool:
    domain = (domain or '').strip().lower()
    if not domain:
        return False

    # Valida resolución DNS básica del dominio de correo.
    import socket

    try:
        socket.getaddrinfo(domain, None)
        return True
    except socket.gaierror:
        return False


def _validate_registration_email(email: str):
    normalized_email = (email or '').strip().lower()
    if not normalized_email:
        raise ValidationError('El correo es obligatorio.')

    try:
        validate_email(normalized_email)
    except ValidationError:
        raise ValidationError('El formato del correo no es válido.')

    domain = normalized_email.split('@')[-1]
    blocked_domains = {'example.com', 'test.com', 'mailinator.com', 'invalid.com'}
    if domain in blocked_domains:
        raise ValidationError('El dominio del correo no está permitido.')

    if not _is_email_domain_reachable(domain):
        raise ValidationError('No se pudo validar el dominio del correo. Verifica que exista.')

    return normalized_email


def _get_role_from_user(user: User) -> str: # pyright: ignore[reportInvalidTypeForm]
    return 'jefe' if user.is_superuser else 'admin' if user.is_staff else 'usuario'


def _profile_existing_fields():
    return {
        field.name
        for field in UserProfile._meta.get_fields()
        if getattr(field, 'concrete', False)
    }


def _safe_update_profile(profile: UserProfile, values: dict):
    existing = _profile_existing_fields()
    update_fields = []
    for key, value in values.items():
        if key in existing:
            setattr(profile, key, value)
            update_fields.append(key)

    if update_fields:
        profile.save(update_fields=update_fields)
    else:
        profile.save()


def _generate_email_verification_code() -> str:
    return ''.join(secrets.choice('0123456789') for _ in range(6))


def _issue_email_verification_code(user: User): # pyright: ignore[reportInvalidTypeForm]
    code = _generate_email_verification_code()
    profile, _ = UserProfile.objects.get_or_create(user=user)
    expires_at = timezone.now() + timedelta(minutes=15)

    _safe_update_profile(
        profile,
        {
            'email_verification_code_hash': make_password(code),
            'email_verification_code_expires_at': expires_at,
            'email_verification_attempts': 0,
            'email_verification_sent_at': timezone.now(),
        },
    )

    if 'email_verification_code_hash' not in _profile_existing_fields():
        return '', None

    return code, expires_at


def _build_verification_link(request, user: User) -> str: # pyright: ignore[reportInvalidTypeForm]
    uid = urlsafe_base64_encode(force_bytes(user.pk))
    token = default_token_generator.make_token(user)
    frontend_base = getattr(settings, 'FRONTEND_APP_URL', '').rstrip('/')
    if frontend_base:
        return f'{frontend_base}/verify-email/{uid}/{token}'

    origin = request.build_absolute_uri('/').rstrip('/')
    return f'{origin}/verify-email/{uid}/{token}'


def _send_verification_email(request, user: User): # pyright: ignore[reportInvalidTypeForm]
    verification_link = _build_verification_link(request, user)
    verification_code, expires_at = _issue_email_verification_code(user)
    code_line = ''
    expiry_line = ''
    if verification_code:
        code_line = f'Codigo de verificacion (expira en 15 minutos): {verification_code}\n\n'
    if expires_at:
        expiry_line = f'Expira el: {expires_at.strftime("%Y-%m-%d %H:%M:%S")}\n\n'

    subject = 'SIGIRL - Verificación de correo'
    message = (
        'Hola,\n\n'
        'Para activar tu cuenta en SIGIRL debes verificar tu correo.\n\n'
        f'{code_line}'
        f'Enlace de verificacion: {verification_link}\n\n'
        f'{expiry_line}'
        'Si no solicitaste este registro, ignora este mensaje.'
    )

    send_mail(
        subject,
        message,
        getattr(settings, 'DEFAULT_FROM_EMAIL', 'no-reply@sigirl.local'),
        [user.email],
        fail_silently=False,
    )

    return {
        'verification_link': verification_link,
        'verification_code': verification_code,
    }


# =====================
# TOKEN (PUBLIC)
# =====================
class PublicTokenObtainPairView(TokenObtainPairView):
    permission_classes = [AllowAny]

    # Sobrescribe el login JWT para devolver también el rol del usuario.
    def post(self, request, *args, **kwargs):
        username = (request.data.get('username') or '').strip()
        password = request.data.get('password') or ''
        user = User.objects.filter(username=username).first()

        if user and user.check_password(password):
            profile, _ = UserProfile.objects.get_or_create(user=user)
            email_verification_required = getattr(settings, 'EMAIL_VERIFICATION_REQUIRED', True)
            profile_email_verified = getattr(profile, 'email_verified', True)
            if email_verification_required and not profile_email_verified:
                return Response(
                    {
                        'error': 'Tu correo no está verificado. Revisa tu bandeja y confirma tu cuenta.',
                        'email_not_verified': True,
                    },
                    status=status.HTTP_403_FORBIDDEN,
                )

        response = super().post(request, *args, **kwargs)

        # --- FIX: asegurar que el campo 'role' siempre se incluya ---
        if response.status_code == 200:
            try:
                user = User.objects.get(username=username)
                role = _get_role_from_user(user)
                # Si response.data es inmutable, crear un nuevo dict
                data = dict(response.data)
                data["role"] = role
                data["username"] = user.username
                response.data = data
            except User.DoesNotExist:
                pass

        return response


class PublicTokenRefreshView(TokenRefreshView):
    permission_classes = [AllowAny]


# =====================
# PRODUCTO (VIEWSET)
# =====================
class ProductoViewSet(viewsets.ModelViewSet):
    queryset = Producto.objects.all()
    serializer_class = ProductoSerializer
    permission_classes = [IsInventoryManagerOrReadOnly]


# =====================
# CATEGORIA
# =====================
class CategoriaViewSet(viewsets.ModelViewSet):
    queryset = Categoria.objects.all()
    serializer_class = CategoriaSerializer
    permission_classes = [IsAuthenticated]


# =====================
# MOVIMIENTO
# =====================
class MovimientoViewSet(viewsets.ModelViewSet):
    queryset = Movimiento.objects.all()
    serializer_class = MovimientoSerializer
    permission_classes = [IsAuthenticated]


# =====================
# PEDIDO
# =====================
class PedidoViewSet(viewsets.ModelViewSet):
    queryset = Pedido.objects.all()
    serializer_class = PedidoSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        queryset = super().get_queryset().select_related('producto', 'usuario')
        if user.is_staff or user.is_superuser:
            return queryset
        return queryset.filter(usuario=user)

    def perform_create(self, serializer):
        user = self.request.user
        profile = getattr(user, 'profile', None)
        serializer.save(
            usuario=user,
            solicitante=user.get_full_name().strip() or user.username,
            creado_por=user.username,
            departamento=getattr(profile, 'department', '') or '',
        )

    def update(self, request, *args, **kwargs):
        pedido = self.get_object()
        user = request.user
        es_dueno = pedido.usuario == user
        puede_editar = user.is_staff or user.is_superuser or (es_dueno and pedido.estado == 'pendiente')
        if not puede_editar:
            return Response({'error': 'No tienes permiso para actualizar este pedido.'}, status=status.HTTP_403_FORBIDDEN)

        payload = request.data.copy()
        estado = str(payload.get('estado', pedido.estado)).lower()
        estado_anterior = pedido.estado

        if estado == 'rechazado' and not str(payload.get('motivo_rechazo', pedido.motivo_rechazo or '')).strip():
            return Response({'motivo_rechazo': ['Debes indicar un motivo de rechazo.']}, status=status.HTTP_400_BAD_REQUEST)

        if estado == 'aprobado' and estado_anterior != 'aprobado':
            producto = pedido.producto
            if producto.cantidad >= pedido.cantidad:
                producto.cantidad -= pedido.cantidad
                producto.save()
                from .models import Movimiento
                Movimiento.objects.create(
                    producto=producto,
                    tipo='salida',
                    cantidad=pedido.cantidad,
                    observacion=f"Pedido aprobado (ID: {pedido.id})"
                )
            else:
                return Response({'error': 'No hay suficiente stock'}, status=status.HTTP_400_BAD_REQUEST)

        payload['estado'] = estado
        if estado in ('aprobado', 'rechazado'):
            payload['fecha_respuesta'] = date.today()

        serializer = self.get_serializer(pedido, data=payload, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()

        # === RF-034: Guardar historial de cambios de estado ===
        if estado != estado_anterior:
            from .models import PedidoHistorial
            PedidoHistorial.objects.create(
                pedido=pedido,
                estado=estado,
                usuario_modificador=user,
                comentario=payload.get('comentario', '')
            )

        return Response(serializer.data)


# =====================
# REGISTER
# =====================
@api_view(['POST'])
@permission_classes([AllowAny])
def register(request):
    username = (request.data.get('username') or '').strip()
    email = (request.data.get('email') or '').strip()
    password = request.data.get('password') or ''
    first_name = (request.data.get('first_name') or '').strip()
    last_name = (request.data.get('last_name') or '').strip()
    requested_role = request.data.get('role', 'usuario')
    institution = (request.data.get('institution') or '').strip()
    department = (request.data.get('department') or '').strip()

    if not username or not password or not email:
        return Response({'error': 'Debes enviar usuario, correo y contraseña.'}, status=status.HTTP_400_BAD_REQUEST)

    if User.objects.filter(username__iexact=username).exists():
        return Response({'username': ['El usuario ya existe.']}, status=status.HTTP_400_BAD_REQUEST)

    if User.objects.filter(email__iexact=email).exists():
        return Response({'email': ['Ese correo ya está registrado.']}, status=status.HTTP_400_BAD_REQUEST)

    try:
        email = _validate_registration_email(email)
    except ValidationError as exc:
        return Response({'email': [str(exc)]}, status=status.HTTP_400_BAD_REQUEST)

    user = User.objects.create_user(
        username=username,
        email=email,
        password=password,
        first_name=first_name,
        last_name=last_name,
    )

    if requested_role == 'admin':
        user.is_staff = True
    elif requested_role in ['jefe', 'jefe_superior']:
        user.is_staff = True
        user.is_superuser = True

    user.save()

    profile, _ = UserProfile.objects.get_or_create(user=user)
    _safe_update_profile(
        profile,
        {
            'institution': institution,
            'department': department,
            'email_verified': False,
            'email_verified_at': None,
        },
    )

    email_verification_required = getattr(settings, 'EMAIL_VERIFICATION_REQUIRED', True)
    verification_meta = {'verification_link': '', 'verification_code': ''}

    if not email_verification_required:
        _safe_update_profile(
            profile,
            {
                'email_verified': True,
                'email_verified_at': timezone.now(),
            },
        )
    else:
        try:
            verification_meta = _send_verification_email(request, user)
        except Exception:
            # En producción puede no haber SMTP listo aún; no bloqueamos creación de cuenta.
            verification_meta = {'verification_link': '', 'verification_code': ''}

    response_payload = {
        'mensaje': 'Usuario creado correctamente.',
        'requires_verification': email_verification_required,
        'username': user.username,
        'email': user.email,
        'role': _get_role_from_user(user),
    }

    if email_verification_required:
        response_payload['mensaje'] = 'Usuario creado. Revisa tu correo para activar la cuenta.'

    if settings.DEBUG:
        response_payload['verification_link'] = verification_meta.get('verification_link', '')
        response_payload['verification_code'] = verification_meta.get('verification_code', '')
        response_payload['verification_code_hint'] = 'En produccion solo se enviara al correo.'

    return Response(response_payload, status=status.HTTP_201_CREATED)


@api_view(['GET'])
@permission_classes([AllowAny])
def verify_email(request, uidb64, token):
    try:
        uid = force_str(urlsafe_base64_decode(uidb64))
        user = User.objects.get(pk=uid)
    except Exception:
        return Response({'error': 'El enlace de verificación no es válido.'}, status=status.HTTP_400_BAD_REQUEST)

    if not default_token_generator.check_token(user, token):
        return Response({'error': 'El enlace de verificación expiró o ya no es válido.'}, status=status.HTTP_400_BAD_REQUEST)

    profile, _ = UserProfile.objects.get_or_create(user=user)
    _safe_update_profile(
        profile,
        {
            'email_verified': True,
            'email_verified_at': timezone.now(),
            'email_verification_code_hash': '',
            'email_verification_code_expires_at': None,
            'email_verification_attempts': 0,
        },
    )

    return Response({'mensaje': 'Correo verificado correctamente. Ya puedes iniciar sesión.'}, status=status.HTTP_200_OK)


@api_view(['POST'])
@permission_classes([AllowAny])
def verify_email_code(request):
    username = (request.data.get('username') or '').strip()
    email = (request.data.get('email') or '').strip().lower()
    code = ''.join(ch for ch in str(request.data.get('code') or '') if ch.isdigit())

    if not code or len(code) != 6:
        return Response({'error': 'Debes ingresar un codigo de 6 digitos.'}, status=status.HTTP_400_BAD_REQUEST)

    user = None
    if username:
        user = User.objects.filter(username__iexact=username).first()
    elif email:
        user = User.objects.filter(email__iexact=email).first()

    if not user:
        return Response({'error': 'No encontramos una cuenta con esos datos.'}, status=status.HTTP_404_NOT_FOUND)

    profile, _ = UserProfile.objects.get_or_create(user=user)
    if getattr(profile, 'email_verified', False):
        return Response({'mensaje': 'La cuenta ya esta verificada.'}, status=status.HTTP_200_OK)

    fields = _profile_existing_fields()
    if 'email_verification_code_hash' not in fields:
        return Response({'error': 'Verificacion por codigo no disponible en este entorno.'}, status=status.HTTP_503_SERVICE_UNAVAILABLE)

    if not getattr(profile, 'email_verification_code_hash', '') or not getattr(profile, 'email_verification_code_expires_at', None):
        return Response({'error': 'Debes solicitar un nuevo codigo de verificacion.'}, status=status.HTTP_400_BAD_REQUEST)

    if profile.email_verification_code_expires_at < timezone.now():
        return Response({'error': 'El codigo expiro. Solicita un nuevo codigo.'}, status=status.HTTP_400_BAD_REQUEST)

    if getattr(profile, 'email_verification_attempts', 0) >= 5:
        return Response({'error': 'Demasiados intentos fallidos. Solicita un nuevo codigo.'}, status=status.HTTP_429_TOO_MANY_REQUESTS)

    if not check_password(code, profile.email_verification_code_hash):
        attempts = getattr(profile, 'email_verification_attempts', 0) + 1
        _safe_update_profile(profile, {'email_verification_attempts': attempts})
        remaining = max(0, 5 - attempts)
        return Response({'error': f'Codigo incorrecto. Intentos restantes: {remaining}.'}, status=status.HTTP_400_BAD_REQUEST)

    _safe_update_profile(
        profile,
        {
            'email_verified': True,
            'email_verified_at': timezone.now(),
            'email_verification_code_hash': '',
            'email_verification_code_expires_at': None,
            'email_verification_attempts': 0,
        },
    )

    return Response({'mensaje': 'Correo verificado correctamente con codigo. Ya puedes iniciar sesion.'}, status=status.HTTP_200_OK)


@api_view(['POST'])
@permission_classes([AllowAny])
def resend_verification_email(request):
    username = (request.data.get('username') or '').strip()
    email = (request.data.get('email') or '').strip().lower()

    user = None
    if username:
        user = User.objects.filter(username__iexact=username).first()
    elif email:
        user = User.objects.filter(email__iexact=email).first()

    if not user:
        return Response({'error': 'No encontramos una cuenta con esos datos.'}, status=status.HTTP_404_NOT_FOUND)

    profile, _ = UserProfile.objects.get_or_create(user=user)
    if getattr(profile, 'email_verified', False):
        return Response({'mensaje': 'La cuenta ya está verificada.'}, status=status.HTTP_200_OK)

    try:
        verification_meta = _send_verification_email(request, user)
    except Exception:
        return Response({'error': 'No fue posible reenviar el correo ahora. Intenta nuevamente.'}, status=status.HTTP_503_SERVICE_UNAVAILABLE)

    response_payload = {'mensaje': 'Se reenviaron las instrucciones de verificación a tu correo.'}
    if settings.DEBUG:
        response_payload['verification_link'] = verification_meta.get('verification_link', '')
        response_payload['verification_code'] = verification_meta.get('verification_code', '')
        response_payload['verification_code_hint'] = 'En produccion el codigo solo se envia por correo.'

    return Response(response_payload, status=status.HTTP_200_OK)


# =====================
# AUTH - Get Current User
# =====================
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_current_user(request):
    user = request.user
    profile, _ = UserProfile.objects.get_or_create(user=user)
    role = _get_role_from_user(user)
    return Response({
        "id": user.id,
        "username": user.username,
        "email": user.email,
        "email_verified": profile.email_verified,
        "is_staff": user.is_staff,
        "is_superuser": user.is_superuser,
        "first_name": user.first_name,
        "last_name": user.last_name,
        "role": role,
    })


# =====================
# AUTH - Manage Profile
# =====================
def _serialize_profile_response(user, profile):
    role = _get_role_from_user(user)
    return {
        'id': user.id,
        'username': user.username,
        'email': user.email,
        'email_verified': profile.email_verified,
        'first_name': user.first_name,
        'last_name': user.last_name,
        'role': role,
        'profile': {
            'institution': profile.institution,
            'department': profile.department,
            'phone': profile.phone,
            'cargo': profile.cargo,
            'bio': profile.bio,
            'avatar': profile.avatar,
        }
    }


@api_view(['GET', 'PATCH', 'DELETE'])
@permission_classes([IsAuthenticated])
def manage_profile(request):
    user = request.user
    profile, _ = UserProfile.objects.get_or_create(user=user)

    if request.method == 'GET':
        return Response(_serialize_profile_response(user, profile))

    if request.method == 'DELETE':
        user.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    data = request.data
    username = str(data.get('username', user.username)).strip()
    email = str(data.get('email', user.email)).strip()

    if username and username != user.username and User.objects.exclude(pk=user.pk).filter(username=username).exists():
        return Response({'username': ['Ya existe un usuario con ese nombre.']}, status=status.HTTP_400_BAD_REQUEST)
    if email and email != user.email and User.objects.exclude(pk=user.pk).filter(email=email).exists():
        return Response({'email': ['Ya existe un usuario con ese correo.']}, status=status.HTTP_400_BAD_REQUEST)

    user.username = username or user.username
    user.first_name = data.get('first_name', user.first_name)
    user.last_name = data.get('last_name', user.last_name)
    user.email = email
    user.save()

    profile_data = data.get('profile', {}) if isinstance(data.get('profile', {}), dict) else {}
    profile.institution = profile_data.get('institution', data.get('institution', profile.institution))
    profile.department = profile_data.get('department', data.get('department', profile.department))
    profile.phone = profile_data.get('phone', data.get('phone', profile.phone))
    profile.cargo = profile_data.get('cargo', data.get('cargo', profile.cargo))
    profile.bio = profile_data.get('bio', data.get('bio', profile.bio))
    profile.avatar = profile_data.get('avatar', data.get('avatar', profile.avatar))
    profile.save()

    return Response(_serialize_profile_response(user, profile))


@api_view(['GET'])
@permission_classes([AllowAny])
def download_inventory_template_excel(request):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Plantilla Inventario'

    headers = ['nombre', 'cantidad', 'categoria', 'fecha']
    sheet.append(headers)
    sheet.append(['Acetona', 25, 'Solventes', '2026-04-27'])
    sheet.append(['Guantes Nitrilo', 100, 'EPP', '2026-04-27'])

    header_fill = PatternFill(start_color='1FA971', end_color='1FA971', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    centered = Alignment(horizontal='center', vertical='center')

    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=column_index)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = centered
        sheet.column_dimensions[chr(64 + column_index)].width = max(14, len(header) + 5)

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.alignment = Alignment(horizontal='left', vertical='center')

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="plantilla_inventario_sigirl.xlsx"'
    return response


@api_view(['GET'])
@permission_classes([AllowAny])
def download_inventory_excel(request):
    workbook = Workbook()
    # openpyxl usa UTF-8 por defecto, pero forzamos encoding en la respuesta
    sheet = workbook.active
    sheet.title = 'Inventario'

    headers = ['Nombre', 'Tipo', 'Categoría', 'Cantidad', 'Umbral Mínimo', 'Ubicación', 'Estado']
    sheet.append(headers)

    productos = Producto.objects.select_related('categoria').order_by('nombre')
    for producto in productos:
        sheet.append([
            producto.nombre,
            producto.tipo,
            producto.categoria.nombre if producto.categoria else '',
            producto.cantidad,
            producto.minimo,
            producto.ubicacion or '',
            producto.estado,
        ])

    header_fill = PatternFill(start_color='0F7A53', end_color='0F7A53', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    for column_index in range(1, len(headers) + 1):
        cell = sheet.cell(row=1, column=column_index)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        sheet.column_dimensions[chr(64 + column_index)].width = 20

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet; charset=utf-8',
    )
    response['Content-Disposition'] = 'attachment; filename="inventario_sigirl.xlsx"'
    response['Content-Encoding'] = 'utf-8'
    return response


def _register_pdf_font():
    import importlib

    pdfmetrics = importlib.import_module('reportlab.pdfbase.pdfmetrics')
    ttfonts = importlib.import_module('reportlab.pdfbase.ttfonts')
    TTFont = getattr(ttfonts, 'TTFont')

    candidate_fonts = [
        ('SegoeUI', 'C:/Windows/Fonts/segoeui.ttf'),
        ('Calibri', 'C:/Windows/Fonts/calibri.ttf'),
    ]
    for font_name, font_path in candidate_fonts:
        try:
            pdfmetrics.registerFont(TTFont(font_name, font_path))
            return font_name
        except Exception:
            continue
    return 'Helvetica'


@api_view(['GET'])
@permission_classes([AllowAny])
def download_inventory_pdf(request):
    import importlib

    colors = importlib.import_module('reportlab.lib.colors')
    A4 = getattr(importlib.import_module('reportlab.lib.pagesizes'), 'A4')
    ParagraphStyle = getattr(importlib.import_module('reportlab.lib.styles'), 'ParagraphStyle')
    mm = getattr(importlib.import_module('reportlab.lib.units'), 'mm')
    platypus = importlib.import_module('reportlab.platypus')
    Paragraph = getattr(platypus, 'Paragraph')
    SimpleDocTemplate = getattr(platypus, 'SimpleDocTemplate')
    Spacer = getattr(platypus, 'Spacer')
    Table = getattr(platypus, 'Table')
    TableStyle = getattr(platypus, 'TableStyle')

    buffer = BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=16 * mm,
        rightMargin=16 * mm,
        topMargin=18 * mm,
        bottomMargin=14 * mm,
    )

    # Usar fuente DejaVuSans para compatibilidad UTF-8
    font_name = _register_pdf_font(font_path=None, font_name='DejaVuSans') if '_register_pdf_font' in globals() else 'Helvetica'
    title_style = ParagraphStyle(
        name='Title',
        fontName=font_name,
        fontSize=16,
        leading=20,
        textColor=colors.HexColor('#0F7A53'),
        spaceAfter=8,
    )
    subtitle_style = ParagraphStyle(
        name='Subtitle',
        fontName=font_name,
        fontSize=9,
        leading=12,
        textColor=colors.HexColor('#4B5563'),
        spaceAfter=12,
    )

    productos = Producto.objects.select_related('categoria').order_by('nombre')
    table_data = [['Producto', 'Categoria', 'Cantidad', 'Minimo', 'Ubicacion', 'Estado']]
    for p in productos:
        table_data.append([
            p.nombre,
            p.categoria.nombre if p.categoria else '',
            str(p.cantidad),
            str(p.minimo),
            p.ubicacion or '-',
            p.estado,
        ])

    table = Table(table_data, repeatRows=1, colWidths=[70 * mm, 34 * mm, 20 * mm, 20 * mm, 30 * mm, 22 * mm])
    table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), font_name),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8.5),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1FA971')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (2, 1), (3, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.HexColor('#F3FBF7')]),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#C6E8D9')),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 7),
        ('TOPPADDING', (0, 0), (-1, 0), 7),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 5),
        ('TOPPADDING', (0, 1), (-1, -1), 5),
    ]))

    story = [
        Paragraph('SIGIRL - Reporte de Inventario', title_style),
        Paragraph(f'Generado: {timezone.localtime().strftime("%Y-%m-%d %H:%M")}', subtitle_style),
        Spacer(1, 4),
        table,
    ]
    document.build(story)

    pdf_data = buffer.getvalue()
    buffer.close()

    response = HttpResponse(pdf_data, content_type='application/pdf; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="reporte_inventario_sigirl.pdf"'
    response['Content-Encoding'] = 'utf-8'
    return response


# =====================
# ALERTA (VIEWSET)
# =====================
class AlertaViewSet(viewsets.ModelViewSet):
    queryset = Alerta.objects.all().order_by('-fecha')
    serializer_class = AlertaSerializer
    permission_classes = [IsAuthenticated]


# =====================
# USER MANAGEMENT (VIEWSET)
# =====================
class UserManagementViewSet(viewsets.ModelViewSet):
    queryset = User.objects.all().order_by('username')
    serializer_class = UserManagementSerializer
    permission_classes = [IsStaffOrSuperuser]

    def create(self, request, *args, **kwargs):
        # Solo admin/jefe pueden crear usuarios y asignar roles
        if not (request.user.is_staff or request.user.is_superuser):
            return Response({'error': 'No tienes permiso para crear usuarios.'}, status=status.HTTP_403_FORBIDDEN)

        data = request.data
        username = str(data.get('username') or data.get('nombre_input') or '').strip()
        email = str(data.get('email') or '').strip()
        password = str(data.get('password') or '').strip()
        rol = data.get('rol_input') or data.get('rol') or 'usuario'

        if not username:
            return Response({'username': ['El nombre de usuario es obligatorio.']}, status=status.HTTP_400_BAD_REQUEST)
        if not password:
            return Response({'password': ['La contraseña es obligatoria.']}, status=status.HTTP_400_BAD_REQUEST)
        if User.objects.filter(username=username).exists():
            return Response({'username': ['Ya existe un usuario con ese nombre.']}, status=status.HTTP_400_BAD_REQUEST)
        if email and User.objects.filter(email=email).exists():
            return Response({'email': ['Ya existe un usuario con ese correo.']}, status=status.HTTP_400_BAD_REQUEST)

        # Admin solo puede crear "admin" o "usuario", no "jefe"
        if request.user.is_staff and not request.user.is_superuser:
            if rol == 'jefe':
                return Response({'rol': ['Solo jefe maestro puede crear otros jefes.']}, status=status.HTTP_403_FORBIDDEN)

        # Jefe maestro puede crear cualquier rol

        user = User(username=username, email=email)
        nombre = str(data.get('nombre_completo') or data.get('full_name') or username).strip()
        parts = nombre.split(' ', 1)
        user.first_name = parts[0]
        user.last_name = parts[1] if len(parts) > 1 else ''

        if rol == 'jefe':
            user.is_staff = True
            user.is_superuser = True
        elif rol == 'admin':
            user.is_staff = True
            user.is_superuser = False
        else:
            user.is_staff = False
            user.is_superuser = False

        user.set_password(password)
        user.save()

        profile, _ = UserProfile.objects.get_or_create(user=user)
        profile.department = str(data.get('departamento_input') or data.get('department') or '').strip()
        profile.save()

        serializer = self.get_serializer(user)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        # Solo admin/jefe pueden editar usuarios
        if not (request.user.is_staff or request.user.is_superuser):
            return Response({'error': 'No tienes permiso para editar usuarios.'}, status=status.HTTP_403_FORBIDDEN)

        user = self.get_object()
        data = request.data
        rol = data.get('rol_input') or data.get('rol') or ''

        # Admin solo puede cambiar rol a "admin" o "usuario", no a "jefe"
        if rol and request.user.is_staff and not request.user.is_superuser:
            if rol == 'jefe':
                return Response({'rol': ['Solo jefe maestro puede asignar rol de jefe.']}, status=status.HTTP_403_FORBIDDEN)

        # Jefe maestro puede cambiar a cualquier rol
        # (no hay restricción adicional)

        username = str(data.get('username') or '').strip()
        if username and username != user.username:
            if User.objects.exclude(pk=user.pk).filter(username=username).exists():
                return Response({'username': ['Ya existe un usuario con ese nombre.']}, status=status.HTTP_400_BAD_REQUEST)
            user.username = username

        nombre = str(data.get('nombre_input') or data.get('nombre_completo') or '').strip()
        if nombre:
            parts = nombre.split(' ', 1)
            user.first_name = parts[0]
            user.last_name = parts[1] if len(parts) > 1 else ''

        email = str(data.get('email') or '').strip()
        if email:
            if email != user.email and User.objects.exclude(pk=user.pk).filter(email=email).exists():
                return Response({'email': ['Ya existe un usuario con ese correo.']}, status=status.HTTP_400_BAD_REQUEST)
            user.email = email

        if rol == 'jefe':
            user.is_staff = True
            user.is_superuser = True
        elif rol == 'admin':
            user.is_staff = True
            user.is_superuser = False
        elif rol == 'usuario':
            user.is_staff = False
            user.is_superuser = False

        password = data.get('password', '')
        if password:
            user.set_password(password)

        user.save()

        profile, _ = UserProfile.objects.get_or_create(user=user)
        profile.department = str(data.get('departamento_input') or data.get('department') or '').strip()
        profile.save()

        serializer = self.get_serializer(user)
        return Response(serializer.data)

    def destroy(self, request, *args, **kwargs):
        user = self.get_object()
        if user == request.user:
            return Response({"error": "No puedes eliminar tu propia cuenta"}, status=400)
        user.delete()
        return Response(status=204)
    
    from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from .serializers import CurrentUserProfileSerializer

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_current_user(request):
    """Obtiene el perfil del usuario actual con su rol"""
    serializer = CurrentUserProfileSerializer(request.user)
    return Response(serializer.data)


from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_current_user(request):
    """Obtiene el perfil del usuario actual con su rol"""
    user = request.user
    return Response({
        'id': user.id,
        'username': user.username,
        'email': user.email,
        'first_name': user.first_name,
        'last_name': user.last_name,
        'role': 'admin' if user.is_superuser else ('jefe' if user.is_staff else 'usuario'),
        'full_name': user.get_full_name() or user.username,
        'is_staff': user.is_staff,
        'is_superuser': user.is_superuser,
    })


# ============================================================
# VIEWSETS PARA GESTIÓN ACADÉMICA
# ============================================================

class ProgramaViewSet(viewsets.ModelViewSet):
    queryset = Programa.objects.all()
    serializer_class = ProgramaSerializer
    
    def get_permissions(self):
        from .permissions import IsAdminOrJefe
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            return [permissions.IsAuthenticated(), IsAdminOrJefe()]
        return [permissions.IsAuthenticated()]

class CompetenciaViewSet(viewsets.ModelViewSet):
    queryset = Competencia.objects.all()
    serializer_class = CompetenciaSerializer
    
    def get_permissions(self):
        from .permissions import IsAdminOrJefe
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            return [permissions.IsAuthenticated(), IsAdminOrJefe()]
        return [permissions.IsAuthenticated()]
    
    def get_queryset(self):
        queryset = Competencia.objects.all()
        programa_id = self.request.query_params.get('programa')
        if programa_id:
            queryset = queryset.filter(programa_id=programa_id)
        return queryset
    
    from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from .models import Practica, PracticaReactivo, PracticaEquipo, Pedido
from django.db import transaction

# ============================================================
# ENDPOINTS PARA CÁLCULO AUTOMÁTICO DE PEDIDOS
# ============================================================

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def calcular_pedido(request):
    print("=== FUNCIÓN CALCULAR_PEDIDO EJECUTADA ===")
    practica_id = request.data.get('practica_id')
    numero_grupos = request.data.get('numero_grupos', 1)
    
    if not practica_id:
        return Response({'error': 'Se requiere practica_id'}, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        practica = Practica.objects.get(id=practica_id)
    except Practica.DoesNotExist:
        return Response({'error': 'Práctica no encontrada'}, status=status.HTTP_404_NOT_FOUND)
    
    # Calcular reactivos
    reactivos = []
    for pr in practica.reactivos.all():
        cantidad_total = float(pr.cantidad) * numero_grupos
        reactivos.append({
            'id': pr.reactivo.id,
            'nombre': pr.reactivo.nombre,
            'cantidad_base': float(pr.cantidad),
            'unidad': pr.unidad.simbolo,
            'cantidad_total': cantidad_total,
            'stock_actual': pr.reactivo.cantidad,
            'suficiente': pr.reactivo.cantidad >= cantidad_total
        })
    
    # Calcular equipos
    equipos = []
    for pe in practica.equipos.all():
        cantidad_total = 1 * numero_grupos
        equipos.append({
            'id': pe.equipo.id,
            'nombre': pe.equipo.nombre,
            'cantidad_base': 1,
            'cantidad_total': cantidad_total,
            'stock_actual': pe.equipo.cantidad,
            'suficiente': pe.equipo.cantidad >= cantidad_total
        })
    
    # ========== CALCULAR MATERIALES ==========
    materiales = []
    for pm in practica.materiales.all():
        cantidad_total = pm.cantidad_por_grupo * numero_grupos
        materiales.append({
            'nombre': pm.nombre,
            'cantidad_por_grupo': pm.cantidad_por_grupo,
            'cantidad_total': cantidad_total
        })
    
    return Response({
        'practica': {
            'id': practica.id,
            'nombre': practica.nombre,
            'ficha': practica.ficha
        },
        'numero_grupos': numero_grupos,
        'reactivos': reactivos,
        'equipos': equipos,
        'materiales': materiales,
        'tiene_stock_suficiente': all(r['suficiente'] for r in reactivos) and all(e['suficiente'] for e in equipos)
    })

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def generar_pedido(request):
    practica_id = request.data.get('practica_id')
    numero_grupos = request.data.get('numero_grupos', 1)
    observaciones = request.data.get('observaciones', '')
    
    try:
        practica = Practica.objects.get(id=practica_id)
    except Practica.DoesNotExist:
        return Response({'error': 'Práctica no encontrada'}, status=404)
    
    pedidos_creados = []
    requiere_aprobacion = False
    
    with transaction.atomic():
        for pr in practica.reactivos.all():
            cantidad_total = float(pr.cantidad) * numero_grupos
            stock_suficiente = pr.reactivo.cantidad >= cantidad_total
            
            pedido = Pedido.objects.create(
                usuario=request.user,
                producto=pr.reactivo,
                cantidad=int(cantidad_total),
                estado='pendiente' if stock_suficiente else 'requiere_aprobacion',
                prioridad='media',
                solicitante=request.user.get_full_name() or request.user.username,
                observaciones=f"Práctica: {practica.nombre}\nGrupos: {numero_grupos}\n{observaciones}",
                requiere_aprobacion_jefe=not stock_suficiente
            )
            pedidos_creados.append(pedido.id)
            if not stock_suficiente:
                requiere_aprobacion = True
    
    return Response({
        'success': True,
        'solicitud_id': practica.id,
        'pedidos_ids': pedidos_creados,
        'requiere_aprobacion': requiere_aprobacion,
        'mensaje': f'Se generaron {len(pedidos_creados)} pedidos. {"Requiere aprobación del Jefe" if requiere_aprobacion else "Todo en stock"}'
    }, status=status.HTTP_201_CREATED)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def aprobar_excepcion_pedido(request, pedido_id):
    user = request.user
    if not (user.is_staff or user.is_superuser):
        return Response({'error': 'Solo el Jefe puede aprobar excepciones'}, status=403)
    
    try:
        pedido = Pedido.objects.get(id=pedido_id)
    except Pedido.DoesNotExist:
        return Response({'error': 'Pedido no encontrado'}, status=404)
    
    if pedido.estado != 'requiere_aprobacion':
        return Response({'error': 'Este pedido no requiere aprobación especial'}, status=400)
    
    pedido.estado = 'aprobado'
    pedido.aprobado_por_jefe = True
    pedido.fecha_aprobacion_jefe = timezone.now()
    pedido.save()
    
    return Response({'success': True, 'mensaje': 'Pedido aprobado excepcionalmente'})

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def pedidos_requieren_aprobacion(request):
    user = request.user
    if not (user.is_staff or user.is_superuser):
        return Response({'error': 'Solo el Jefe puede ver esta lista'}, status=status.HTTP_403_FORBIDDEN)
    
    pedidos = Pedido.objects.filter(estado='requiere_aprobacion').select_related('producto', 'usuario')
    
    data = []
    for p in pedidos:
        data.append({
            'id': p.id,
            'codigo': p.codigo,
            'producto': p.producto.nombre,
            'cantidad': p.cantidad,
            'solicitante': p.solicitante,
            'fecha_solicitud': p.fecha_solicitud,
            'observaciones': p.observaciones,
            'stock_actual': p.producto.cantidad
        })
    
    return Response(data)


# ============================================================
# GENERAR PDF DEL FORMATO DE SOLICITUD
# ============================================================

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm, cm
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from io import BytesIO
from datetime import datetime

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def generar_pdf_solicitud(request):
    """Genera PDF con el formato de solicitud de laboratorio"""
    
    practica_id = request.data.get('practica_id')
    numero_grupos = request.data.get('numero_grupos', 1)
    observaciones = request.data.get('observaciones', '')
    
    try:
        practica = Practica.objects.get(id=practica_id)
    except Practica.DoesNotExist:
        return Response({'error': 'Práctica no encontrada'}, status=404)
    
    # Calcular cantidades
    reactivos = []
    for pr in practica.reactivos.all():
        cantidad_total = float(pr.cantidad) * numero_grupos
        reactivos.append({
            'nombre': pr.reactivo.nombre,
            'cantidad_total': cantidad_total,
            'unidad': pr.unidad.simbolo
        })
    
    materiales = []
    for pm in practica.materiales.all():
        cantidad_total = pm.cantidad_por_grupo * numero_grupos
        materiales.append({
            'nombre': pm.nombre,
            'cantidad_por_grupo': pm.cantidad_por_grupo,
            'cantidad_total': cantidad_total
        })
    
    equipos = []
    for pe in practica.equipos.all():
        cantidad_total = 1 * numero_grupos
        equipos.append({
            'nombre': pe.equipo.nombre,
            'cantidad_total': cantidad_total
        })
    
    # Crear PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm, leftMargin=1.5*cm, rightMargin=1.5*cm)
    styles = getSampleStyleSheet()
    
    # Estilo de título
    title_style = ParagraphStyle(
        'TitleStyle',
        parent=styles['Heading1'],
        fontSize=12,
        alignment=1,  # Centro
        spaceAfter=10,
        fontName='Helvetica-Bold'
    )
    
    subtitle_style = ParagraphStyle(
        'SubtitleStyle',
        parent=styles['Normal'],
        fontSize=9,
        alignment=1,
        spaceAfter=15
    )
    
    header_style = ParagraphStyle(
        'HeaderStyle',
        parent=styles['Normal'],
        fontSize=9,
        alignment=1,
        fontName='Helvetica-Bold'
    )
    
    cell_style = ParagraphStyle(
        'CellStyle',
        parent=styles['Normal'],
        fontSize=8
    )
    
    elements = []
    
    # Título principal
    elements.append(Paragraph("FORMATO SOLICITUD DE REACTIVOS, MATERIALES, EQUIPOS E INSTRUMENTOS", title_style))
    elements.append(Paragraph("LABORATORIOS DE QUÍMICA Y BIOTECNOLOGÍA v.0.3", subtitle_style))
    elements.append(Spacer(1, 5))
    
    # Información general
    info_data = [
        ["Laboratorio No.", "______", "GRUPO", "______", "Especialidad", "______"],
        ["Fecha elaboración", datetime.now().strftime('%Y-%m-%d'), "Fecha de la práctica", practica.fecha.strftime('%Y-%m-%d') if practica.fecha else "______", "Hora de la práctica", "______"],
        ["Número de grupos", str(numero_grupos), "Competencia", practica.competencia.nombre if practica.competencia else "______", "Instructor", request.user.username],
        ["Nombre de la práctica", practica.nombre, "", "", "", ""],
    ]
    
    for row in info_data:
        t = Table([row], colWidths=[3*cm, 3*cm, 3*cm, 3*cm, 3*cm, 3*cm])
        t.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(t)
        elements.append(Spacer(1, 5))
    
    elements.append(Spacer(1, 10))
    
    # Tabla de Reactivos
    elements.append(Paragraph("SOLICITUD DE REACTIVOS Y SOLUCIONES", header_style))
    elements.append(Spacer(1, 5))
    
    reactivo_headers = ["REACTIVO", "CANTIDAD REQUERIDA (g/ml)", "PESO INICIAL (g)", "PESO FINAL (g)", "CANTIDAD CONSUMIDA (g)", "NOMBRE SOLUCION", "CONCENTRACION", "CANTIDAD REQUERIDA (ml)"]
    reactivo_data = [reactivo_headers]
    for r in reactivos:
        reactivo_data.append([r['nombre'], f"{r['cantidad_total']}", "___", "___", "___", "___", "___", "___"])
    
    t_reactivos = Table(reactivo_data, colWidths=[3*cm, 2.5*cm, 2*cm, 2*cm, 2.5*cm, 2.5*cm, 2.5*cm, 2.5*cm])
    t_reactivos.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(t_reactivos)
    elements.append(Spacer(1, 10))
    
    # Tabla de Materiales
    elements.append(Paragraph("MATERIALES", header_style))
    elements.append(Spacer(1, 5))
    
    material_headers = ["ELEMENTO", "CANTIDAD POR GRUPO", "CANTIDAD TOTAL", "ELEMENTO", "CANTIDAD POR GRUPO", "CANTIDAD TOTAL"]
    material_data = [material_headers]
    
    # Organizar materiales en 2 columnas
    half = len(materiales) // 2 + len(materiales) % 2
    col1 = materiales[:half]
    col2 = materiales[half:]
    
    for i in range(max(len(col1), len(col2))):
        row = []
        if i < len(col1):
            row.extend([col1[i]['nombre'], str(col1[i]['cantidad_por_grupo']), str(col1[i]['cantidad_total'])])
        else:
            row.extend(["", "", ""])
        if i < len(col2):
            row.extend([col2[i]['nombre'], str(col2[i]['cantidad_por_grupo']), str(col2[i]['cantidad_total'])])
        else:
            row.extend(["", "", ""])
        material_data.append(row)
    
    t_materiales = Table(material_data, colWidths=[4*cm, 2.5*cm, 2.5*cm, 4*cm, 2.5*cm, 2.5*cm])
    t_materiales.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(t_materiales)
    elements.append(Spacer(1, 10))
    
    # Tabla de Equipos
    elements.append(Paragraph("EQUIPOS / INSTRUMENTOS", header_style))
    elements.append(Spacer(1, 5))
    
    equipos_data = [["EQUIPO", "CANTIDAD", "ACCESORIOS", "CONDICIONES DE ALISTAMIENTO Y OPERACIÓN"]]
    for e in equipos:
        equipos_data.append([e['nombre'], str(e['cantidad_total']), "___", "___"])
    
    t_equipos = Table(equipos_data, colWidths=[5*cm, 2.5*cm, 3*cm, 5*cm])
    t_equipos.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(t_equipos)
    elements.append(Spacer(1, 10))
    
    # Observaciones
    elements.append(Paragraph(f"OBSERVACIONES: {observaciones or 'Ninguna'}", cell_style))
    elements.append(Spacer(1, 15))
    
    # Firmas
    firmas_data = [
        ["LOGÍSTICA", "APRENDIZ ENCARGADO", "OBSERVACIONES", "FIRMA RECIBIDO DE MATERIALES DE FORMACION"],
        ["CABINAS", "_________________", "_________________", "_________________"],
        ["REACTIVOS", "", "", ""],
        ["MATERIAL", "", "NOMBRE INSTRUCTOR", "FIRMA INSTRUCTOR"],
        ["BALANZAS", "", "_________________", "_________________"],
        ["MESONES", "", "", ""],
        ["PISO", "", "NOMBRE VOCERO", "FIRMA VOCERO"],
        ["POCETAS", "", "_________________", "_________________"],
        ["", "", "", ""],
        ["", "", "NOMBRE TECNICO", "FIRMA TECNICO"],
        ["", "", "_________________", "_________________"],
        ["FIRMA LÍDER LABORATORIO", "_________________", "", ""],
    ]
    
    t_firmas = Table(firmas_data, colWidths=[3.5*cm, 4*cm, 4*cm, 4*cm])
    t_firmas.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(t_firmas)
    
    # Construir PDF
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    
    # Devolver PDF
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="solicitud_{practica.nombre.replace(" ", "_")}.pdf"'
    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def reporte_sustancias_controladas(request):
    from datetime import date
    
    user = request.user
    if not (user.is_staff or user.is_superuser):
        return Response({'error': 'Solo administradores pueden ver este reporte'}, status=403)
    
    # Mostrar TODOS los reactivos, no solo los sensibles
    reactivos = Producto.objects.filter(tipo='reactivo')
    
    data = []
    for r in reactivos:
        alertas = []
        if r.cantidad <= r.minimo:
            alertas.append(f"Stock bajo: {r.cantidad} (mínimo: {r.minimo})")
        if r.fecha_vencimiento:
            dias = (r.fecha_vencimiento - date.today()).days
            if dias <= 30:
                alertas.append(f"Vence en {dias} días")
        
        data.append({
            'id': r.id,
            'nombre': r.nombre,
            'cantidad': r.cantidad,
            'minimo': r.minimo,
            'ubicacion': r.ubicacion,
            'fecha_vencimiento': r.fecha_vencimiento,
            'es_sensible': r.es_sensible,
            'alertas': alertas
        })
    
    return Response({
        'total_reactivos': len(data),
        'total_sensibles': len([r for r in data if r['es_sensible']]),
        'reactivos_con_alerta': len([r for r in data if r['alertas']]),
        'reactivos': data
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def toggle_reactivo_sensible(request, reactivo_id):
    """Activar/desactivar la marca de reactivo sensible"""
    user = request.user
    if not (user.is_staff or user.is_superuser):
        return Response({'error': 'Solo administradores pueden hacer esto'}, status=403)
    
    try:
        reactivo = Producto.objects.get(id=reactivo_id, tipo='reactivo')
        reactivo.es_sensible = not reactivo.es_sensible
        reactivo.save()
        
        return Response({
            'success': True,
            'id': reactivo.id,
            'nombre': reactivo.nombre,
            'es_sensible': reactivo.es_sensible,
            'mensaje': f'Reactivo {"marcado como sensible" if reactivo.es_sensible else "desmarcado como sensible"}'
        })
    except Producto.DoesNotExist:
        return Response({'error': 'Reactivo no encontrado'}, status=404)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def crear_reactivo_sensible(request):
    """Crear un nuevo reactivo y marcarlo como sensible"""
    user = request.user
    if not (user.is_staff or user.is_superuser):
        return Response({'error': 'Solo administradores pueden hacer esto'}, status=403)
    
    data = request.data
    nombre = data.get('nombre')
    cantidad = data.get('cantidad', 0)
    minimo = data.get('minimo', 10)
    ubicacion = data.get('ubicacion', '')
    categoria_nombre = data.get('categoria', 'Reactivos')
    
    if not nombre:
        return Response({'error': 'El nombre es obligatorio'}, status=400)
    
    # Obtener o crear categoría
    categoria, _ = Categoria.objects.get_or_create(nombre=categoria_nombre)
    
    reactivo = Producto.objects.create(
        nombre=nombre,
        tipo='reactivo',
        categoria=categoria,
        cantidad=int(cantidad),
        minimo=int(minimo),
        ubicacion=ubicacion,
        unidad='ml',
        es_sensible=True  # Marcar como sensible
    )
    
    return Response({
        'success': True,
        'id': reactivo.id,
        'nombre': reactivo.nombre,
        'es_sensible': reactivo.es_sensible,
        'mensaje': f'Reactivo sensible "{nombre}" creado exitosamente'
    }, status=201)